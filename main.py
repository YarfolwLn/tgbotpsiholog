#Подключение библиотек
#pip install aiogram python-docx openpyxl

#Содержимое файла requirements.txt
#aiogram==3.13.0
#aiohttp==3.9.5
#openpyxl==3.1.5

#Установка зависимостей
#pip install -r requirements.txt

import asyncio
import logging
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import os
from datetime import datetime, timedelta
from openpyxl.styles import Alignment

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Токен бота
BOT_TOKEN = os.getenv("TELEGRAM_TOKEN")

# Инициализация бота и диспетчера
bot = Bot(token=BOT_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)

# Путь к файлу Excel
EXCEL_FILE = "appointments.xlsx"

# Создаем файл Excel если его нет
def init_excel_file():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Записи"
        
        # Заголовки (измененные)
        headers = ["Дни недели", "Время", "Имя пользователя", "Телеграм ID", "Телефон", "Ситуация", "Статус"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Устанавливаем ширину колонок
        column_widths = [20, 20, 20, 15, 15, 30, 15]
        for col, width in enumerate(column_widths, 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[col_letter].width = width
        
        # Создаем стили для ячеек
        text_style = NamedStyle(name="text_style", number_format='@')  # Текстовый формат
        
        # Добавляем стили в книгу
        if 'text_style' not in wb.named_styles:
            wb.add_named_style(text_style)
        
        # Устанавливаем форматы для столбцов
        for row in range(2, 100):  # Устанавливаем для большего количества строк
            for col in range(1, 8):  # Все колонки текстовые
                ws.cell(row=row, column=col).style = 'text_style'
        
        wb.save(EXCEL_FILE)
        logger.info("Excel файл создан с правильными форматами ячеек")

# Состояния FSM
class AppointmentState(StatesGroup):
    user_name = State()
    user_phone = State()
    user_situation = State()
    choosing_days = State()
    entering_time_for_days = State()

# Класс для работы с Excel
class ExcelManager:
    def __init__(self, file_path):
        self.file_path = file_path
        self.red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    def get_next_empty_row(self, ws):
        """Находит следующую пустую строку в таблице"""
        row = 2  # Начинаем с 2 строки (после заголовков)
        while ws.cell(row=row, column=1).value is not None:
            row += 1
        return row
    
    def book_appointment(self, days_str, time_range_str, username, user_id, phone, situation):
        """Записываем данные с днями недели и диапазоном времени"""
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            
            # Находим первую свободную строку
            new_row = self.get_next_empty_row(ws)
            
            # Записываем данные (измененные колонки)
            ws.cell(row=new_row, column=1, value=str(days_str))      # Дни недели
            ws.cell(row=new_row, column=2, value=str(time_range_str)) # Диапазон времени
            ws.cell(row=new_row, column=3, value=str(username))      # Имя
            ws.cell(row=new_row, column=4, value=str(user_id))       # Телеграм ID
            ws.cell(row=new_row, column=5, value=str(phone))         # Телефон
            ws.cell(row=new_row, column=6, value=str(situation))     # Ситуация
            ws.cell(row=new_row, column=7, value="Ожидает подтверждения")  # Статус
            
            # Красим строку для визуального выделения
            for col in range(1, 8):
                ws.cell(row=new_row, column=col).fill = self.red_fill
            
            wb.save(self.file_path)
            wb.close()
            logger.info(f"Запись сохранена в строке {new_row}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при записи в Excel: {e}")
            return False
    
    def book_multiple_appointments(self, selected_days, days_with_times, username, user_id, phone, situation):
        """Создает отдельные записи для каждой пары день-время"""
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            
            success_count = 0
            
            # Для каждого дня создаем отдельную запись
            for day in selected_days:
                time_range = days_with_times.get(day, "")
                if time_range:
                    # Находим первую свободную строку для каждой записи
                    new_row = self.get_next_empty_row(ws)
                    
                    # Записываем данные для каждого дня отдельно
                    ws.cell(row=new_row, column=1, value=str(day))          # День недели
                    ws.cell(row=new_row, column=2, value=str(time_range))   # Диапазон времени
                    ws.cell(row=new_row, column=3, value=str(username))     # Имя
                    ws.cell(row=new_row, column=4, value=str(user_id))      # Телеграм ID
                    ws.cell(row=new_row, column=5, value=str(phone))        # Телефон
                    ws.cell(row=new_row, column=6, value=str(situation))    # Ситуация
                    ws.cell(row=new_row, column=7, value="Ожидает подтверждения")  # Статус
                    
                    # Красим строку для визуального выделения
                    for col in range(1, 8):
                        ws.cell(row=new_row, column=col).fill = self.red_fill
                    
                    success_count += 1
                    logger.info(f"Запись для дня {day} сохранена в строке {new_row}")
            
            wb.save(self.file_path)
            wb.close()
            logger.info(f"Создано {success_count} записей в Excel")
            return success_count > 0
            
        except Exception as e:
            logger.error(f"Ошибка при записи в Excel: {e}")
            return False

# Инициализация менеджера Excel
excel_manager = ExcelManager(EXCEL_FILE)

# Клавиатуры
def get_main_keyboard():
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📅 Записаться на прием")],
            [KeyboardButton(text="🆘 Помощь")]
        ],
        resize_keyboard=True,
        input_field_placeholder="Выберите действие..."
    )
    return keyboard

def get_back_to_main_keyboard():
    """Клавиатура с кнопкой возврата в главное меню"""
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="↩️ В главное меню")]],
        resize_keyboard=True,
        one_time_keyboard=True
    )

def get_days_keyboard():
    """Клавиатура для выбора дней недели с кнопкой возврата в главное меню"""
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Понедельник"), KeyboardButton(text="Вторник")],
            [KeyboardButton(text="Среда"), KeyboardButton(text="Четверг")],
            [KeyboardButton(text="Пятница"), KeyboardButton(text="Суббота")],
            [KeyboardButton(text="Воскресенье")],
            [KeyboardButton(text="✅ Завершить выбор дней")],
            [KeyboardButton(text="↩️ В главное меню")]
        ],
        resize_keyboard=True,
        one_time_keyboard=True
    )
    return keyboard

def get_time_input_keyboard():
    """Клавиатура для ввода времени с кнопкой возврата в главное меню"""
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="↩️ В главное меню")]],
        resize_keyboard=True,
        one_time_keyboard=True
    )

# Функция для проверки корректности времени
def is_valid_time_range(time_str):
    """Проверяет корректность формата диапазона времени (например, 9:00-12:00)"""
    try:
        if '-' not in time_str:
            return False, "❌ Используйте формат ЧЧ:MM-ЧЧ:MM (например, 9:00-12:00)"
        
        start_time_str, end_time_str = time_str.split('-')
        
        # Проверяем оба времени
        start_time = datetime.strptime(start_time_str.strip(), '%H:%M')
        end_time = datetime.strptime(end_time_str.strip(), '%H:%M')
        
        if start_time >= end_time:
            return False, "❌ Время начала должно быть раньше времени окончания"
        
        return True, "Диапазон времени корректен"
        
    except ValueError:
        return False, "❌ Неверный формат времени. Используйте ЧЧ:MM-ЧЧ:MM (например, 9:00-12:00):"

# Функция для отправки уведомлений администратору
async def send_notification_to_admin(user_data, days_with_times):
    """Отправка уведомления администратору о новой заявке"""
    try:
        admin_chat_id = os.getenv("ADMIN_ID")  # Замените на нужный ID администратора
        
        notification_text = (
            "🔔 НОВАЯ ЗАЯВКА НА КОНСУЛЬТАЦИЮ\n\n"
            f"👤 Имя клиента: {user_data['user_name']}\n"
            f"📞 Телефон: {user_data['user_phone']}\n"
            f"🆔 Telegram ID: {user_data['user_id']}\n"
        )
        
        if user_data['user_situation']:
            notification_text += f"📝 Ситуация: {user_data['user_situation']}\n"
        
        notification_text += f"\n📅 Предпочтительные дни и время:\n"
        for day, time_range in days_with_times.items():
            notification_text += f"• {day}: {time_range}\n"
        
        notification_text += "\n📋 Статус: Ожидает подтверждения\n"
        notification_text += "\n⚠️ Свяжитесь с клиентом для подтверждения записи"
        
        await bot.send_message(chat_id=admin_chat_id, text=notification_text)
        logger.info(f"Уведомление отправлено администратору о новой заявке")
        
    except Exception as e:
        logger.error(f"Ошибка при отправке уведомления администратору: {e}")

# Обработчики команд
@dp.message(Command("start"))
async def cmd_start(message: types.Message, state: FSMContext):
    # Очищаем состояние, если пользователь был в процессе записи
    current_state = await state.get_state()
    if current_state:
        await state.clear()
        await message.answer(
            "❌ Процесс записи прерван. Возвращаемся в главное меню.",
            reply_markup=get_main_keyboard()
        )
    else:
        await message.answer(
            "👋 Добро пожаловать в бота по записи на приём!\n\n"
            "Я ваш виртуальный помощник. Я могу\n"
            "📅 Записать вас на прием к психологу\n\n"
            "Выберите действие из меню ниже:",
            reply_markup=get_main_keyboard()
        )

@dp.message(Command("help"))
async def cmd_help(message: types.Message, state: FSMContext):
    # Очищаем состояние, если пользователь был в процессе записи
    current_state = await state.get_state()
    if current_state:
        await state.clear()
        await message.answer(
            "❌ Процесс записи прерван. Возвращаемся в главное меню.",
            reply_markup=get_main_keyboard()
        )
    
    help_text = """
🆘 Помощь по боту:

📅 Запись на прием:
- Нажмите «📅 Записаться на прием»
- Введите ваше имя
- Введите ваш номер телефона
- Опишите вашу ситуацию
- Выберите подходящие дни недели
- Для каждого дня введите удобный диапазон времени

⚙️ Управление:
- ↩️ В главное меню» - вернуться в главное меню

Для начала работы нажмите /start
    """
    await message.answer(help_text, reply_markup=get_main_keyboard())

@dp.message(F.text == "🆘 Помощь")
async def help_command(message: types.Message, state: FSMContext):
    await cmd_help(message, state)

@dp.message(F.text == "📅 Записаться на прием")
async def book_appointment(message: types.Message, state: FSMContext):
    await message.answer(
        "👤 Введите ваше имя:\n\n",
        reply_markup=get_back_to_main_keyboard()
    )
    await state.set_state(AppointmentState.user_name)

# Обработка имени
@dp.message(AppointmentState.user_name)
async def process_name(message: types.Message, state: FSMContext):
    if message.text == "↩️ В главное меню":
        await back_to_main_process(message, state)
        return
        
    if len(message.text.strip()) < 2:
        await message.answer(
            "❌ Имя должно содержать хотя бы 2 символа. Пожалуйста, введите ваше имя:\n\n",
            reply_markup=get_back_to_main_keyboard()
        )
        return
        
    await state.update_data(user_name=message.text.strip())
    await message.answer(
        "📞 Теперь введите ваш номер телефона:\n\n",
        reply_markup=get_back_to_main_keyboard()
    )
    await state.set_state(AppointmentState.user_phone)

# Обработка телефона
@dp.message(AppointmentState.user_phone)
async def process_phone(message: types.Message, state: FSMContext):
    if message.text == "↩️ В главное меню":
        await back_to_main_process(message, state)
        return
        
    phone = message.text.strip()
    if len(phone) < 5:
        await message.answer(
            "❌ Номер телефона слишком короткий. Пожалуйста, введите корректный номер:\n\n",
            reply_markup=get_back_to_main_keyboard()
        )
        return
        
    await state.update_data(user_phone=phone)
    await message.answer(
        "📝 Опишите кратко вашу ситуацию или проблему, с которой хотите обратиться "
        "(это поможет психологу лучше подготовиться к встрече):\n\n"
        "Если не хотите описывать, отправьте \"-\" или \"пропустить\"\n\n",
        reply_markup=get_back_to_main_keyboard()
    )
    await state.set_state(AppointmentState.user_situation)

# Обработка ситуации
@dp.message(AppointmentState.user_situation)
async def process_situation(message: types.Message, state: FSMContext):
    if message.text == "↩️ В главное меню":
        await back_to_main_process(message, state)
        return
        
    situation = message.text.strip()
    if situation.lower() in ["-", "пропустить", "нет", "не хочу"]:
        situation = ""
        
    await state.update_data(user_situation=situation)
    
    await message.answer(
        "📅 Теперь выберите подходящие дни недели для приема:\n\n"
        "Нажимайте на кнопки с днями недели, которые вам подходят.\n"
        "Вы можете выбрать несколько дней.\n"
        "Если хотите удалить день из списка - нажмите на него повторно.\n"
        "Когда закончите, нажмите «✅ Завершить выбор дней»\n\n",
        reply_markup=get_days_keyboard()
    )
    await state.set_state(AppointmentState.choosing_days)
    await state.update_data(selected_days=[])  # Инициализируем пустой список выбранных дней

# Обработка выбора дней недели
@dp.message(AppointmentState.choosing_days)
async def process_days_selection(message: types.Message, state: FSMContext):
    if message.text == "↩️ В главное меню":
        await back_to_main_process(message, state)
        return
        
    user_data = await state.get_data()
    selected_days = user_data.get('selected_days', [])
    
    days_of_week = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
    
    if message.text in days_of_week:
        if message.text not in selected_days:
            selected_days.append(message.text)
            await message.answer(
                f"✅ День добавлен: {message.text}\n\n"
                f"Выбранные дни: {', '.join(selected_days)}\n\n"
                f"Продолжайте выбирать дни или нажмите «✅ Завершить выбор дней»",
                reply_markup=get_days_keyboard()
            )
        else:
            selected_days.remove(message.text)
            await message.answer(
                f"❌ Удален: {message.text}\n\n"
                f"Выбранные дни: {', '.join(selected_days)}\n\n"
                f"Продолжайте выбирать дни или нажмите «✅ Завершить выбор дней»",
                reply_markup=get_days_keyboard()
            )
        
        await state.update_data(selected_days=selected_days)
        
    elif message.text == "✅ Завершить выбор дней":
        if not selected_days:
            await message.answer(
                "❌ Вы не выбрали ни одного дня. Пожалуйста, выберите хотя бы один день:\n\n",
                reply_markup=get_days_keyboard()
            )
            return
        
        # Сохраняем выбранные дни и начинаем ввод времени для каждого дня
        await state.update_data(selected_days=selected_days)
        await state.update_data(days_with_times={})  # Словарь для хранения времени по дням
        await state.update_data(current_day_index=0)  # Индекс текущего дня
        
        user_data = await state.get_data()
        selected_days = user_data['selected_days']
        
        # Начинаем с первого дня
        first_day = selected_days[0]
        await message.answer(
            f"✅ Выбраны дни: {', '.join(selected_days)}\n\n"
            f"⏰ Теперь введите удобное время для выбранных дней в формате ЧЧ:MM-ЧЧ:MM\n"
            "Например: 9:00-12:00 или 14:00-16:00\n\n"
            f"{first_day}:",
            reply_markup=get_time_input_keyboard()
        )
        await state.set_state(AppointmentState.entering_time_for_days)
        
    else:
        await message.answer(
            "Пожалуйста, выберите дни недели из предложенных вариантов:\n\n",
            reply_markup=get_days_keyboard()
        )

# Обработка ввода времени для каждого дня
@dp.message(AppointmentState.entering_time_for_days)
async def process_time_for_days(message: types.Message, state: FSMContext):
    if message.text == "↩️ В главное меню":
        await back_to_main_process(message, state)
        return
        
    user_data = await state.get_data()
    selected_days = user_data.get('selected_days', [])
    days_with_times = user_data.get('days_with_times', {})
    current_day_index = user_data.get('current_day_index', 0)
    
    current_day = selected_days[current_day_index]
    
    # Проверяем корректность введенного времени
    is_valid, message_text = is_valid_time_range(message.text.strip())
    
    if not is_valid:
        await message.answer(
            f"{message_text}\n\n"
            f"Введите время для выбранного дня\n\n"
            f"{current_day}:",
            reply_markup=get_time_input_keyboard()
        )
        return
    
    # Сохраняем время для текущего дня
    days_with_times[current_day] = message.text.strip()
    await state.update_data(days_with_times=days_with_times)
    
    # Переходим к следующему дню
    next_day_index = current_day_index + 1
    
    if next_day_index < len(selected_days):
        # Есть еще дни для ввода времени
        next_day = selected_days[next_day_index]
        await state.update_data(current_day_index=next_day_index)
        
        await message.answer(
            f"✅ День недели: {current_day}, Время: {message.text.strip()}\n\n"
            f"⏰ Теперь введите удобное время для следующего выбранного дня в формате ЧЧ:MM-ЧЧ:MM\n\n"
            "Например: 9:00-12:00 или 14:00-16:00\n\n"
            f"{next_day}:",
            reply_markup=get_time_input_keyboard()
        )
    else:
        # Все дни обработаны, завершаем запись
        user_name = user_data['user_name']
        user_phone = user_data['user_phone']
        user_situation = user_data.get('user_situation', '')
        user_id = message.from_user.id
        
        # Создаем отдельные записи для каждого дня
        success = excel_manager.book_multiple_appointments(
            selected_days, days_with_times, user_name, user_id, user_phone, user_situation
        )
        
        if success:
            response = (
                f"✅ Заявка успешно отправлена!\n\n"
                f"👤 Имя: {user_name}\n"
                f"📞 Телефон: {user_phone}\n"
            )
            
            if user_situation:
                response += f"📝 Ситуация: {user_situation}\n\n"
            else:
                response += "\n"
                
            response += f"📅 Выбранные дни и время:\n"
            for day, time_range in days_with_times.items():
                response += f"• {day}: {time_range}\n"
            
            response += (
                "\n📞 С вами свяжутся в ближайшее время для уточнения деталей.\n"
            )
            
            await message.answer(response, reply_markup=get_main_keyboard())
            
            # Отправляем уведомление администратору
            user_data['user_id'] = user_id
            await send_notification_to_admin(user_data, days_with_times)
            
        else:
            await message.answer(
                "❌ Произошла ошибка при отправке заявки. Пожалуйста, попробуйте позже.",
                reply_markup=get_main_keyboard()
            )
        
        await state.clear()

@dp.message(F.text == "↩️ В главное меню")
async def back_to_main_process(message: types.Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state:
        await state.clear()
        await message.answer(
            "❌ Процесс записи прерван. Возвращаемся в главное меню.",
            reply_markup=get_main_keyboard()
        )
    else:
        await message.answer(
            "Вы уже в главном меню.",
            reply_markup=get_main_keyboard()
        )

# Обработка обычных сообщений
@dp.message()
async def handle_other_messages(message: types.Message, state: FSMContext):
    if message.text.startswith('/'):
        return
        
    current_state = await state.get_state()
    if current_state:
        # Если пользователь в процессе записи, но ввел что-то не то
        await message.answer(
            "Пожалуйста, следуйте инструкциям процесса записи или нажмите «↩️ В главное меню» для отмены.",
            reply_markup=get_back_to_main_keyboard()
        )
    else:
        await message.answer(
            "Я не совсем понимаю, что вы имеете в виду. "
            "Пожалуйста, используйте кнопки меню или команды для взаимодействия с ботом.",
            reply_markup=get_main_keyboard()
        )

# Основная функция
async def main():
    # Инициализируем Excel файл
    init_excel_file()
    
    logger.info("Бот запущен и готов к работе!")
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())



